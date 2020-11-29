from openpyxl import load_workbook
import csv

PRODUCT_SKU = 0
PRODUCT_QTY = 1
PRODUCT_NAME = 2


class Product:
    def __init__(self, sku, qty, name):
        self.sku = sku
        self.qty = qty
        self.name = name

    def __repr__(self):
        return "{} : {}".format(self.sku, self.qty)


class InventoryProduct(Product):
    def __init__(self, sku, qty):
        self.sku = sku
        self.qty = qty


def update_inventory(inv_list, ven_list):
    result = dict()
    present_sku = []

    for i_p in inv_list:
        for v_p in ven_list:
            if i_p.sku == v_p.sku:
                i_p.qty = v_p.qty
                result.update({i_p: v_p.qty})
                present_sku.append(i_p.sku)
        # discontinued
        if i_p.sku not in present_sku:
            result.update({i_p: '-50'})

    return result


inventroy_file = 'shopify_inventory/corcoran_ca_export_1.csv'
workbook = load_workbook(
    filename='vendor_report/Specs sheet-inventory Corcoran 26 NOV 2020.xlsx', read_only=True)

inv_products, products = [], []

for sheet in workbook:
    for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
        if str(row[PRODUCT_QTY]) != 'None':
            product = Product(row[PRODUCT_SKU],
                              row[PRODUCT_QTY], row[PRODUCT_NAME])
            products.append(product)

with open(inventroy_file, 'r', encoding='utf8') as inventory:
    reader = csv.DictReader(inventory)

    for line in reader:
        product = InventoryProduct(
            line['SKU'], line['Wholesale Furniture Brokers'])
        inv_products.append(product)

inventory.close()
result = update_inventory(inv_products, products)
